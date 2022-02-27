from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.touch_actions import TouchActions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from string import digits
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
#from fake_useragent import UserAgent
from selenium.common import exceptions
import bs4
import re
from selenium.webdriver.support.expected_conditions import staleness_of
import time
from openpyxl import Workbook
import requests
from urllib.request import Request, urlopen
import pandas as pd


def configure_driver():
    # Add additional Options to the webdriver
    chrome_options = Options()
    # ua = UserAgent()
    # userAgent = ua.random  # THIS IS FAKE AGENT IT WILL GIVE YOU NEW AGENT EVERYTIME
    # print(userAgent)
    # add the argument and make the browser Headless.
    # chrome_options.add_argument("--headless")  # if you don't want to see the display on chrome just uncomment this
    # chrome_options.add_argument(f'user-agent={userAgent}')  # useragent added
    chrome_options.add_argument("--log-level=3")  # removes error/warning/info messages displayed on the console
    chrome_options.add_argument("--disable-notifications")  # disable notifications
    chrome_options.add_argument(
        "--disable-infobars")  # disable infobars ""Chrome is being controlled by automated test software"  Although is isn't supported by Chrome anymore
    chrome_options.add_argument("start-maximized")  # will maximize chrome screen
    # chrome_options.add_argument('--disable-gpu')  # disable gpu (not load pictures fully)
    chrome_options.add_argument("--disable-extensions")  # will disable developer mode extensions
    # chrome_options.add_argument('--blink-settings=imagesEnabled=false')
    # chrome_options.add_argument('--proxy-server=%s' % PROXY)
    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    # prefs = {"profile.managed_default_content_settings.images": 2}
    # chrome_options.add_experimental_option("prefs", prefs)             #we have disabled pictures (so no time is wasted in loading them)

    driver = webdriver.Chrome(ChromeDriverManager().install(),
                              options=chrome_options)  # you don't have to download chromedriver it will be downloaded by itself and will be saved in cache
    return driver


def RunScrapper(driver):

    val = ''
    comp1 = "None"
    val_new = ''
    num_list = []
    #   READING THE CATEGORIES  #
    frCAT = open("./List of words.txt", "r", encoding='utf-8')
    catLi = frCAT.read().split("\n")
    frCAT.close()
    for cat in catLi:
        if cat:
            start_time = time.time()
            print("                  ****************************WELCOME TO 2GIS****************************")

            # workbook created
            wb = Workbook()
            # add_sheet is used to create sheet.
            sheet1 = wb.active
            print(" WORKSHEET CREATED SUCCESSFULLY!")
            # INITIALIZING THE COLOUMN NAMES NOW
            c1 = sheet1.cell(row=1, column=1)
            c1.value = "Company Name"
            c2 = sheet1.cell(row=1, column=2)
            c2.value = "Categories"
            c3 = sheet1.cell(row=1, column=3)
            c3.value = "Address"
            c4 = sheet1.cell(row=1, column=4)
            c4.value = "Email"
            c5 = sheet1.cell(row=1, column=5)
            c5.value = "Phone Number"
            c6 = sheet1.cell(row=1, column=6)
            c6.value = "Mobile Number"
            c7 = sheet1.cell(row=1, column=7)
            c7.value = "Land-line Number"
            c8 = sheet1.cell(row=1, column=8)
            c8.value = "Website Link"
            c9 = sheet1.cell(row=1, column=9)
            c9.value = "Location Link"
            wb.save(cat + ".xlsx")

            driver.get("https://2gis.ae/")
            # WebDriverWait(driver, 20).until(
            #     EC.visibility_of_element_located((By.ID, "acceptRiskButton"))).click()
            # input to the search box
            search_box = driver.find_element_by_xpath("//input[@type = 'text']")
            search_box.clear()
            search_box.send_keys(cat)
            search_box.send_keys(Keys.RETURN)
            # click on the FILTERS button
            try:
                WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, "//span[@class = '_1y2da3ul']")))
                driver.find_element_by_xpath("//span[@class = '_1y2da3ul']").click()
            except Exception:
                pass
            count = 1
            mi = 2
            for container in range(1):
                # click on the categories
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, "//div[@class = '_1hf7139']")))
                button_checking = 0
                time.sleep(3)
                while True:
                    names = driver.find_elements_by_xpath("//div[@class = '_1hf7139']")
                    actions = ActionChains(driver)
                    actions.move_to_element(names[-1]).perform()
                    button_checking = button_checking + 1
                    if button_checking == 15:
                        break

                print(len(names))
                WebDriverWait(driver, 20).until(
                    EC.visibility_of_element_located((By.XPATH, "//div[@class = '_1h3cgic']")))
                for name in names:
                    print("*****************", count, "*****************")
                    try:
                        # click on the category
                        category = name.find_element_by_xpath(".//div[@class = '_1h3cgic']")
                        actions = ActionChains(driver)
                        actions.click(category).perform()
                        # Company Name
                        soup = bs4.BeautifulSoup(driver.page_source, 'html.parser')
                        company_Name1 = soup.find("div", class_="_1dcp9fc")  # find all the elements with class _1dcp9fc
                        company_Name = company_Name1.find("h1")
                        # Category
                        category = company_Name1.find("div")
                        # address
                        try:
                            add = driver.find_elements_by_xpath("//div[@class ='_49kxlr']//div[@class = '_1p8iqzw']")
                            print("length:", len(add))
                            if len(add) == 2:
                                address = add[1].text
                            elif len(add) == 1:
                                address = add[0].text
                                print(type(address))
                        except Exception:
                            address = ""

                        # location_link
                        try:
                            location_links = soup.find_all("span", class_="_er2xx9")
                            for location_lin in location_links:
                                try:
                                    location_link = location_lin.find("a").get("href")
                                    if bool(location_link) == True:
                                        location_link = "https://2gis.ae" + location_link
                                        break
                                    elif "http://redirect.2gis.com/account/" or "http://redirect.2gis.com/adv/" in location_link:
                                        location_link = ""
                                        pass
                                except Exception:
                                    pass
                        except Exception:
                            pass

                        # phone and mobile numbers (also with print) and save in xlsx file
                        try:
                            WebDriverWait(driver, 3).until(
                                EC.element_to_be_clickable((By.XPATH, "//div[@class = '_b0ke8']//child :: button")))
                            driver.find_element_by_xpath("//div[@class = '_b0ke8']//child :: button").click()
                        except Exception:
                            pass
                        try:
                            x = 1
                            x_new = 1
                            numbers = driver.find_elements_by_xpath("//div[@class = '_b0ke8']")
                            print("length of phone numbers:", len(numbers))
                            for number in numbers:
                                if "+" in number.text:
                                    mobile_no = number.text
                                    mobile_no = re.sub('[^0-9]', '', mobile_no)
                                    if "9715" in mobile_no:
                                        if x == 1:
                                           val = mobile_no
                                           c6 = sheet1.cell(row=mi, column=6)
                                           c6.value = val
                                           print("Mobile No.:", val)
                                        else:
                                            c1 = sheet1.cell(row=mi, column=6)
                                            value_in_excel = c1.value
                                            val = value_in_excel + "," + mobile_no
                                            c1 = sheet1.cell(row=mi, column=6)
                                            c1.value = val
                                            print("Mobile No.:", val)
                                        x += 1
                                    elif "9714" in mobile_no:
                                        if x_new == 1:
                                            val_new = mobile_no
                                            c1 = sheet1.cell(row=mi, column=7)
                                            c1.value = val_new
                                            print("Land-line No.:", val_new)
                                        else:
                                            c1 = sheet1.cell(row=mi, column=7)
                                            value_in_excel = c1.value
                                            val_new = value_in_excel + "," + mobile_no
                                            c1 = sheet1.cell(row=mi, column=7)
                                            c1.value = val_new
                                            print("Land-line No.:", val_new)
                                        x_new += 1
                                else:
                                    phone_number = number.text
                                    phone_number = re.sub('[^0-9]', '', phone_number)
                                    print("Phone No.:", phone_number)
                                    c5 = sheet1.cell(row=mi, column=5)
                                    c5.value = phone_number
                        except Exception:
                            pass

                        # Get Websites link with print and also save in xlsx file
                        try:
                            website_links = driver.find_elements_by_xpath(
                                "//div[@class='_49kxlr']//a[@href and (@target = '_blank') and (@class = '_1rehek')]")
                            for website_link in website_links:
                                print("Website:", website_link.get_attribute("href"))
                                c8 = sheet1.cell(row=mi, column=8)
                                c8.value = website_link.get_attribute("href")
                        except Exception:
                            pass

                        # Get Emails with print and also save in xlsx file
                        try:
                            containers = driver.find_elements_by_xpath(
                                "//div[@class='_49kxlr']//a[@href  and (@target = '_blank') and (@class = '_2lcm958')]")
                            for check_mail in containers:
                                mail = check_mail.get_attribute("href")
                                if "tel:" in mail:
                                    pass
                                else:
                                    mail = mail.replace("mailto:", "")
                                    print("Mail:", mail)
                                    c4 = sheet1.cell(row=mi, column=4)
                                    c4.value = mail
                        except Exception:
                            pass

                        # print company name with save in xlsx file also
                        print("company_Name:", company_Name.text)
                        c1 = sheet1.cell(row=mi, column=1)
                        c1.value = company_Name.text
                        # print category with save in xlsx file also
                        print("Category:", category.text)
                        c2 = sheet1.cell(row=mi, column=2)
                        c2.value = category.text
                        # print address with save in xlsx file also
                        try:
                            print("Address:", address)
                            c3 = sheet1.cell(row=mi, column=3)
                            c3.value = address
                        except Exception:
                            pass
                        # print location_link with save in xlsx file also
                        try:
                            print("Location Link:", location_link)
                            c9 = sheet1.cell(row=mi, column=9)
                            c9.value = location_link
                        except Exception:
                            pass

                        count = count + 1
                        # save the xlsx file
                        wb.save(cat + ".xlsx")
                        mi = mi + 1
                    except Exception:
                        print("link broken")
                        pass
                # click on the next page
                try:
                    button_checking = 0
                    br = 0
                    while br != 2:
                        names = driver.find_elements_by_xpath("//a[@class = '_12164l30']")
                        actions = ActionChains(driver)
                        actions.move_to_element(names[-1]).perform()
                        button_checking = button_checking + 1
                        if button_checking == 5:
                            br = 2
                except Exception:
                    pass
                try:
                    check_the_button = driver.find_element_by_xpath("//div[@class = '_n5hmn94'] /*[name()='svg']")
                    check_the_button.click()
                    for container in range(1000):
                        # find the names, categories
                        WebDriverWait(driver, 10).until(
                            EC.visibility_of_element_located((By.XPATH, "//div[@class = '_1hf7139']")))
                        time.sleep(2)
                        names = driver.find_elements_by_xpath("//div[@class = '_1hf7139']")
                        print(len(names))

                        for name in names:
                            print("*****************", count, "*****************")
                            try:
                                # click on the category
                                category = name.find_element_by_xpath(".//div[@class = '_1h3cgic']")
                                actions = ActionChains(driver)
                                actions.click(category).perform()
                                # Company Name
                                soup = bs4.BeautifulSoup(driver.page_source, 'html.parser')
                                company_Name1 = soup.find("div",
                                                          class_="_1dcp9fc")  # find all the elements with class _1dcp9fc
                                company_Name = company_Name1.find("h1")
                                # Category
                                category = company_Name1.find("div")
                                # address
                                try:
                                    add = driver.find_elements_by_xpath(
                                        "//div[@class ='_49kxlr']//div[@class = '_1p8iqzw']")
                                    print("length:", len(add))
                                    if len(add) == 2:
                                        address = add[1].text
                                    elif len(add) == 1:
                                        address = add[0].text
                                        print(type(address))
                                except Exception:
                                    address = ""

                                # location_link
                                try:
                                    location_links = soup.find_all("span", class_="_er2xx9")
                                    for location_lin in location_links:
                                        try:
                                            location_link = location_lin.find("a").get("href")
                                            if bool(location_link) == True:
                                                location_link = "https://2gis.ae" + location_link
                                                break
                                            elif "http://redirect.2gis.com/account/" or "http://redirect.2gis.com/adv/" in location_link:
                                                location_link = ""
                                                pass
                                        except Exception:
                                            pass
                                except Exception:
                                    pass

                                # phone and mobile numbers (also with print) and save in xlsx file
                                try:
                                    WebDriverWait(driver, 3).until(
                                        EC.element_to_be_clickable(
                                            (By.XPATH, "//div[@class = '_b0ke8']//child :: button")))
                                    driver.find_element_by_xpath("//div[@class = '_b0ke8']//child :: button").click()
                                except Exception:
                                    pass
                                try:
                                    x = 1
                                    x_new = 1
                                    numbers = driver.find_elements_by_xpath("//div[@class = '_b0ke8']")
                                    print("length of phone numbers:", len(numbers))
                                    for number in numbers:
                                        if "+" in number.text:
                                            mobile_no = number.text
                                            mobile_no = re.sub('[^0-9]', '', mobile_no)
                                            if "9715" in mobile_no:
                                                if x == 1:
                                                    val = mobile_no
                                                    c6 = sheet1.cell(row=mi, column=6)
                                                    c6.value = val
                                                    print("Mobile No.:", val)
                                                else:
                                                    c1 = sheet1.cell(row=mi, column=6)
                                                    value_in_excel = c1.value
                                                    val = value_in_excel + "," + mobile_no
                                                    c1 = sheet1.cell(row=mi, column=6)
                                                    c1.value = val
                                                    print("Mobile No.:", val)
                                                x += 1
                                            elif "9714" in mobile_no:
                                                if x_new == 1:
                                                    val_new = mobile_no
                                                    c1 = sheet1.cell(row=mi, column=7)
                                                    c1.value = val_new
                                                    print("Land-line No.:", val_new)
                                                else:
                                                    c1 = sheet1.cell(row=mi, column=7)
                                                    value_in_excel = c1.value
                                                    val_new = value_in_excel + "," + mobile_no
                                                    c1 = sheet1.cell(row=mi, column=7)
                                                    c1.value = val_new
                                                    print("Land-line No.:", val_new)
                                                x_new += 1
                                        else:
                                            phone_number = number.text
                                            phone_number = re.sub('[^0-9]', '', phone_number)
                                            print("Phone No.:", phone_number)
                                            c5 = sheet1.cell(row=mi, column=5)
                                            c5.value = phone_number
                                except Exception:
                                    pass

                                # Get Websites link with print and also save in xlsx file
                                try:
                                    website_links = driver.find_elements_by_xpath(
                                        "//div[@class='_49kxlr']//a[@href and (@target = '_blank') and (@class = '_1rehek')]")
                                    for website_link in website_links:
                                        print("Website:", website_link.get_attribute("href"))
                                        c8 = sheet1.cell(row=mi, column=8)
                                        c8.value = website_link.get_attribute("href")
                                except Exception:
                                    pass

                                # Get Emails with print and also save in xlsx file
                                try:
                                    containers = driver.find_elements_by_xpath(
                                        "//div[@class='_49kxlr']//a[@href  and (@target = '_blank') and (@class = '_2lcm958')]")
                                    for check_mail in containers:
                                        mail = check_mail.get_attribute("href")
                                        if "tel:" in mail:
                                            pass
                                        else:
                                            mail = mail.replace("mailto:", "")
                                            print("Mail:", mail)
                                            c4 = sheet1.cell(row=mi, column=4)
                                            c4.value = mail
                                except Exception:
                                    pass

                                # print company name with save in xlsx file also
                                print("company_Name:", company_Name.text)
                                c1 = sheet1.cell(row=mi, column=1)
                                c1.value = company_Name.text
                                # print category with save in xlsx file also
                                print("Category:", category.text)
                                c2 = sheet1.cell(row=mi, column=2)
                                c2.value = category.text
                                # print address with save in xlsx file also
                                try:
                                    print("Address:", address)
                                    c3 = sheet1.cell(row=mi, column=3)
                                    c3.value = address
                                except Exception:
                                    pass
                                # print location_link with save in xlsx file also
                                try:
                                    print("Location Link:", location_link)
                                    c9 = sheet1.cell(row=mi, column=9)
                                    c9.value = location_link
                                except Exception:
                                    pass

                                count = count + 1
                                # save the xlsx file
                                wb.save(cat + ".xlsx")
                                mi = mi + 1
                            except Exception:
                                print("link broken")
                                pass

                        # click on the next button
                        time.sleep(0.5)
                        button_checking = 0
                        br = 0
                        while br != 2:
                            names = driver.find_elements_by_xpath("//a[@class = '_12164l30']")
                            actions = ActionChains(driver)
                            actions.move_to_element(names[-1]).perform()
                            button_checking = button_checking + 1
                            if button_checking == 5:
                                br = 2
                        try:
                            check_the_button = driver.find_elements_by_xpath(
                                "//div[@class = '_n5hmn94'] /*[name()='svg']")
                            check_the_button[1].click()
                        except Exception:
                            print("completed")
                            break
                except Exception:
                    print("completed")
                    break






# create the driver object.
driver = configure_driver()
# call the scrapper to run
RunScrapper(driver)
# close the driver.
# driver.close()
